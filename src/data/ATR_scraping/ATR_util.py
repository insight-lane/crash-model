import re
import openpyxl
import matplotlib.pyplot as pyplot

MAP_FP = 'data/processed/maps'


def is_readable_ATR(fname):
    """
    Function to check if ATR is of type we want to read
    checking for 3 conditions

     1) of 'XXX' type (contains speed, volume, and classification data)
     2) of 24-HOURS type
     3) .XLSX file type 
    """

    # split file name so we can check relevant info
    meta_info = fname.split('_')
    file_type = meta_info[8].split('.')[1]

    # only look at files that are .xlsx, are over 24 hours, and are of type XXX
    if (meta_info[7] == 'XXX') and (meta_info[6] == '24-HOURS') and (file_type == 'XLSX'):
        return True
    else:
        return False


def clean_ATR_fname(fname):
    """
    Clean filename to prepare for geocoding
    EX:
    7362_NA_NA_147_TRAIN-ST_DORCHESTER_24-HOURS_XXX_03-19-2014.XLSX
    to
    147 TRAIN ST Boston, MA
    """

    atr_address = fname.split('_') # split address on underscore character
    atr_address = ' '.join(atr_address[3:5]) # combine elements that make up the address
    atr_address = re.sub('-', ' ', atr_address) # replace '-' with spaces
    atr_address += ' Boston, MA'
    return atr_address


def plot_hourly_rates(files, outfile):
    """
    Function that reads ATRs and generates a sparkline plot
    of percentages of traffic over time
    
    Args:
        files - list of filenames to process
        outfile - where to write the resulting plot
    """

    all_counts = []
    for f in files:
        wb = openpyxl.load_workbook(f, data_only=True)
        sheet_names = wb.sheetnames
        if 'Classification-Combined' in sheet_names:
            sheet = wb['Classification-Combined']
            # Right now the cell locations are hardcoded,
            # but if we expand to cover different formats, will need to change
            counts = []
            for row_index in range(9, 33):
                cell = "{}{}".format('O', row_index)
                val = sheet[cell].value
                counts.append(float(val))
            total = sheet['O34'].value
            for i in range(len(counts)):
                counts[i] = counts[i]/total
            all_counts.append(counts)

    bins = list(range(0, 24))
    for val in all_counts:
        pyplot.plot(bins, val)
    pyplot.legend(loc='upper right')
    pyplot.savefig(outfile)


def read_ATR(fname):
    """
    Function to read ATR data
    data to collect:
    mean speed, volume, motos (# of motorcycles), light(# of cars/trucks), 
    and heavy(# of heavy duty vehicles)
    """

    # data_only=True so as to not read formulas
    wb = openpyxl.load_workbook(fname, data_only=True)
    sheet_names = wb.sheetnames

    # get total volume cell F106
    if 'Volume' in sheet_names:
        sheet = wb['Volume']
        vol = sheet['F106'].value
    else:
        vol = 0

    # get mean speed data
    if 'Speed Combined' in sheet_names:
        sheet = wb['Speed Combined']
        speed = sheet['E42'].value
    elif 'Speed-1' in sheet_names:
        sheet = wb['Speed-1']
        speed = sheet['E42'].value
    else:
        speed = 0

    # get classification data
    if 'Classification-Combined' in sheet_names:
        sheet = wb['Classification-Combined']
        motos = sheet['D38'].value
        light = sheet['D39'].value
        heavy = sheet['D40'].value
    elif 'Classification-1' in sheet_names:
        sheet = wb['Classification-1']
        motos = sheet['D38'].value
        light = sheet['D39'].value
        heavy = sheet['D40'].value
    else:
        motos = 0
        light = 0
        heavy = 0

    return vol, speed, motos, light, heavy



